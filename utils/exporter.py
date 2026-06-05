"""Structured Excel reports for transaction evaluation."""

from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Any, Iterable, Mapping

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

VAT_WARNING = (
    "Domestic sale VAT must still be shown on the domestic invoice unless the "
    "product is VAT-exempt, even when import/export VAT credits offset the payable amount."
)

DEFAULT_LABELS = {
    "import_value_foreign": "Import value (foreign currency)",
    "currency": "Currency",
    "currency_to_usd": "Currency to USD conversion factor",
    "import_value_usd": "Import value equivalent (USD)",
    "fx_rate": "FX rate to Toman",
    "hs_code": "HS code",
    "product_description": "Product description",
    "country_of_origin": "Country of origin",
    "quantity_weight": "Quantity / weight",
    "vat_exempt": "VAT-exempt product",
    "special_permits": "Special permits required",
    "registration_owner": "Import registration owner",
    "fx_source_type": "FX source type",
    "expected_export_obligation_settlement": "Expected export obligation settlement",
    "foreign_goods_cost": "Foreign goods cost",
    "freight": "Foreign freight",
    "insurance": "Foreign insurance",
    "fx_transfer_fee_percent": "FX transfer fee percent",
    "fx_spread_percent": "FX spread percent",
    "customs_valuation_adjustment_percent": "Customs valuation adjustment percent",
    "customs_taxable_base_override": "Confirmed customs taxable base override",
    "customs_duty_percent": "Customs duty percent",
    "commercial_benefit_duty_percent": "Commercial benefit duty percent",
    "other_import_duties": "Other import duties",
    "clearance_cost": "Clearance cost",
    "warehousing_cost": "Warehousing cost",
    "demurrage_cost": "Demurrage cost",
    "port_costs": "Port costs",
    "miscellaneous_costs": "Miscellaneous costs",
    "import_vat_rate": "Import VAT rate",
    "domestic_invoice_amount": "Domestic invoice amount",
    "domestic_vat_rate": "Domestic VAT rate",
    "existing_export_vat_credit": "Existing export VAT credit",
    "buyer_pays_vat_separately": "Buyer pays VAT separately",
    "import_vat_payer": "Import VAT payer",
    "invoice_in_taxpayer_system": "E-invoice issued in taxpayer system",
    "buyer_confirmed_invoice": "Buyer confirmed e-invoice",
    "revenue_recognition_model": "Revenue recognition model",
    "commission_per_usd": "Commission per USD",
    "fixed_commission": "Fixed commission",
    "export_obligation_benefit": "Export obligation benefit",
    "inta_code": "Inta code",
    "activity_description": "Inta activity description",
    "estimated_profit_ratio": "Estimated Inta-code profit ratio",
    "inta_profile_aligned": "Activity aligned with card owner's tax profile",
    "tax_advisor_confirmed": "Tax advisor confirmation",
    "tax_rate": "Income tax rate",
    "admin_cost": "Accounting / admin cost",
    "risk_reserve": "Risk reserve deducted from card owner benefit",
    "tax_scenario": "Active income tax scenario",
    "custom_profit_amount": "Custom profit amount",
    "customs_risk_reserve_percent": "Customs risk reserve percent",
    "vat_risk_reserve_percent": "VAT risk reserve percent",
    "income_tax_risk_reserve_percent": "Income tax risk reserve percent",
    "fx_risk_reserve_percent": "FX risk reserve percent",
    "export_obligation_risk_reserve_percent": "Export obligation risk reserve percent",
    "fixed_legal_admin_reserve": "Fixed legal / admin reserve",
    "goods_cost_toman": "Goods cost (Toman)",
    "freight_toman": "Freight (Toman)",
    "insurance_toman": "Insurance (Toman)",
    "foreign_total_toman": "Foreign cost total (Toman)",
    "fx_transfer_fee": "FX transfer fee",
    "fx_spread_cost": "FX spread cost",
    "calculated_customs_taxable_base": "Calculated customs taxable base",
    "customs_taxable_base_override": "Confirmed customs taxable base override",
    "customs_taxable_base": "Customs taxable base used",
    "customs_duty": "Customs duty",
    "commercial_benefit_duty": "Commercial benefit duty",
    "import_vat_base": "Import VAT base",
    "import_vat": "Import VAT",
    "local_costs": "Local import costs",
    "landed_cost_excluding_recoverable_vat": "Landed cost excluding recoverable VAT",
    "total_landed_cost": "Total landed cost",
    "domestic_sale_vat": "Domestic sale VAT shown on invoice",
    "import_vat_credit": "Import VAT credit",
    "net_vat_payable": "Net VAT payable",
    "net_vat_refundable": "Net VAT refundable / credit",
    "net_vat_position": "Net VAT position",
    "gross_commission": "Gross card owner commission",
    "tax_scenario_a": "Income tax scenario A - commission only",
    "tax_scenario_b": "Income tax scenario B - Inta deemed profit",
    "tax_scenario_c": "Income tax scenario C - custom profit",
    "selected_tax_scenario": "Selected income tax scenario",
    "selected_income_tax": "Selected income tax",
    "net_card_owner_benefit": "Net card owner benefit",
    "customs_risk_reserve": "Customs risk reserve",
    "vat_risk_reserve": "VAT risk reserve",
    "income_tax_risk_reserve": "Income tax risk reserve",
    "fx_risk_reserve": "FX risk reserve",
    "export_obligation_risk_reserve": "Export obligation risk reserve",
    "fixed_legal_admin_reserve": "Fixed legal / admin reserve",
    "minimum_collateral": "Minimum collateral required",
    "total_operator_cost": "Total operator / importer cost",
    "risk_score": "Risk score",
    "risk_level": "Risk level",
    "final_decision": "Final decision",
}

SHEET_KEYS = {
    "Customs & Import Costs": [
        "goods_cost_toman", "freight_toman", "insurance_toman", "foreign_total_toman",
        "fx_transfer_fee", "fx_spread_cost", "calculated_customs_taxable_base",
        "customs_taxable_base_override", "customs_taxable_base", "customs_duty",
        "commercial_benefit_duty", "other_import_duties", "local_costs",
        "landed_cost_excluding_recoverable_vat", "total_landed_cost",
    ],
    "VAT Calculation": [
        "customs_taxable_base", "import_vat_base", "import_vat",
        "domestic_invoice_amount", "domestic_sale_vat", "import_vat_credit",
        "existing_export_vat_credit", "net_vat_payable", "net_vat_refundable",
        "net_vat_position",
    ],
    "Collateral Calculator": [
        "customs_risk_reserve", "vat_risk_reserve", "income_tax_risk_reserve",
        "fx_risk_reserve", "export_obligation_risk_reserve",
        "fixed_legal_admin_reserve", "minimum_collateral",
    ],
    "Summary Decision": [
        "gross_commission", "import_vat", "domestic_sale_vat", "net_vat_position",
        "net_vat_payable", "net_vat_refundable",
        "tax_scenario_a", "tax_scenario_b", "tax_scenario_c",
        "selected_tax_scenario", "selected_income_tax", "net_card_owner_benefit",
        "total_operator_cost", "minimum_collateral", "risk_score", "risk_level",
        "final_decision",
    ],
}


def _frame(
    data: Mapping[str, Any],
    labels: Mapping[str, str] | None = None,
    keys: Iterable[str] | None = None,
) -> pd.DataFrame:
    labels = {**DEFAULT_LABELS, **(labels or {})}
    selected_keys = list(keys) if keys is not None else list(data)
    rows = []
    for key in selected_keys:
        value = data.get(key, "")
        if isinstance(value, (dict, list, tuple, set)):
            continue
        rows.append({"Item": labels.get(key, key), "Value": value})
    return pd.DataFrame(rows)


def checklist_frame(checklist: list[Mapping[str, Any]]) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "کنترل": item["label_fa"],
                "انجام شده؟ بله/خیر": "بله" if item.get("checked") else "خیر",
                "وزن ریسک": item["weight"],
                "ریسک لحاظ‌شده": 0 if item.get("checked") else item["weight"],
                "توضیح": item.get("explanation_fa", ""),
            }
            for item in checklist
        ]
    )


def income_tax_frame(outputs: Mapping[str, Any]) -> pd.DataFrame:
    selected = str(outputs.get("selected_tax_scenario", "A"))
    return pd.DataFrame(
        [
            {
                "Scenario": "A - Commission only",
                "Tax Base": outputs.get("commission_tax_base", 0),
                "Income Tax": outputs.get("tax_scenario_a", 0),
                "Active": "Yes" if selected == "A" else "No",
            },
            {
                "Scenario": "B - Deemed profit / Inta code",
                "Tax Base": outputs.get("deemed_profit", 0),
                "Income Tax": outputs.get("tax_scenario_b", 0),
                "Active": "Yes" if selected == "B" else "No",
            },
            {
                "Scenario": "C - Custom profit",
                "Tax Base": outputs.get("custom_profit", 0),
                "Income Tax": outputs.get("tax_scenario_c", 0),
                "Active": "Yes" if selected == "C" else "No",
            },
        ]
    )


def _style_workbook(buffer: BytesIO) -> BytesIO:
    buffer.seek(0)
    workbook = load_workbook(buffer)
    header_fill = PatternFill("solid", fgColor="173F5F")
    warning_fill = PatternFill("solid", fgColor="FFF2CC")
    alert_fill = PatternFill("solid", fgColor="F4CCCC")
    for sheet in workbook.worksheets:
        sheet.sheet_view.rightToLeft = True
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
        for row in sheet.iter_rows(min_row=2):
            row_text = " ".join(str(cell.value or "") for cell in row)
            for cell in row:
                cell.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
                if isinstance(cell.value, (float, int)) and not isinstance(cell.value, bool):
                    cell.number_format = '#,##0.00'
                if "WARNING" in row_text or "Domestic sale VAT must" in row_text:
                    cell.fill = warning_fill
                    cell.font = Font(color="9C6500", bold=True)
                if "Reject" in row_text or "رد معامله" in row_text:
                    cell.fill = alert_fill
                    cell.font = Font(color="9C0006", bold=True)
        for column in sheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column)
            sheet.column_dimensions[column[0].column_letter].width = min(max(max_length + 3, 14), 58)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def export_all_to_excel(
    inputs: Mapping[str, Any],
    outputs: Mapping[str, Any],
    checklist: list[Mapping[str, Any]],
    labels: Mapping[str, str] | None = None,
) -> BytesIO:
    workbook = Workbook()
    workbook.remove(workbook.active)

    def add_sheet(title: str, header: list[str], rows: list[list[Any]]) -> None:
        sheet = workbook.create_sheet(title)
        sheet.append(header)
        for row in rows:
            sheet.append(row)

    add_sheet("ورودی‌های معامله", ["عنوان", "مقدار", "توضیح"], [
        ["ارزش واردات ارزی", inputs.get("import_value_foreign"), ""],
        ["نوع ارز", inputs.get("currency"), ""],
        ["نرخ تبدیل", inputs.get("fx_rate"), "تومان"],
        ["ارزش ریالی واردات", outputs.get("import_value_total"), ""],
        ["HS Code", inputs.get("hs_code"), ""],
        ["شرح کالا", inputs.get("product_description"), ""],
        ["کشور مبدأ", inputs.get("country_of_origin"), ""],
        ["مبلغ فاکتور فروش داخلی قبل از VAT", inputs.get("domestic_invoice_amount"), ""],
        ["کارمزد هر دلار برای صاحب کارت", inputs.get("commission_per_usd"), ""],
        ["میزان منفعت رفع تعهد ارزی", inputs.get("export_obligation_benefit"), ""],
        ["آیا کالا معاف از VAT است؟", "بله" if inputs.get("vat_exempt") else "خیر", ""],
        ["آیا مجوز خاص نیاز دارد؟", "بله" if inputs.get("special_permits") else "خیر", ""],
        ["مبلغ مورد انتظار رفع تعهد", inputs.get("expected_export_obligation_settlement"), ""],
        ["آیا VAT فروش داخلی جداگانه توسط خریدار پرداخت می‌شود؟", "بله" if inputs.get("buyer_pays_vat_separately") else "خیر", ""],
    ])

    add_sheet("هزینه‌های گمرکی و واردات", ["شرح", "مبنا/نرخ", "مبلغ تومان"], [
        ["قیمت کالا", inputs.get("foreign_goods_cost"), outputs.get("goods_cost_toman")],
        ["حمل", inputs.get("freight"), outputs.get("freight_toman")],
        ["بیمه", inputs.get("insurance"), outputs.get("insurance_toman")],
        ["کارمزد حواله %", inputs.get("fx_transfer_fee_percent"), outputs.get("fx_transfer_fee")],
        ["اسپرد تبدیل ارز %", inputs.get("fx_spread_percent"), outputs.get("fx_spread_cost")],
        ["تعدیل ارزش گمرکی %", inputs.get("customs_valuation_adjustment_percent"), outputs.get("customs_taxable_base")],
        ["پایه ارزش گمرکی", "", outputs.get("customs_taxable_base")],
        ["حقوق ورودی %", inputs.get("customs_duty_percent"), outputs.get("customs_duty")],
        ["سود بازرگانی %", inputs.get("commercial_benefit_duty_percent"), outputs.get("commercial_benefit_duty")],
        ["سایر عوارض", "", outputs.get("other_import_duties")],
        ["هزینه ترخیص", "", inputs.get("clearance_cost")],
        ["انبارداری", "", inputs.get("warehousing_cost")],
        ["دموراژ", "", inputs.get("demurrage_cost")],
        ["هزینه‌های بندری", "", inputs.get("port_costs")],
        ["بهای تمام‌شده قبل از VAT", "", outputs.get("landed_cost_excluding_recoverable_vat")],
        ["بهای تمام‌شده با VAT", "", outputs.get("total_landed_cost")],
    ])

    add_sheet("محاسبات ارزش افزوده", ["شرح", "نرخ/ورودی", "مبلغ تومان", "توضیح"], [
        ["پایه مشمول VAT واردات", "", outputs.get("import_vat_base"), ""],
        ["نرخ VAT واردات", inputs.get("import_vat_rate"), "", ""],
        ["مبلغ VAT واردات", "", outputs.get("import_vat"), ""],
        ["مبلغ فاکتور فروش داخلی", "", inputs.get("domestic_invoice_amount"), ""],
        ["نرخ VAT فروش داخلی", inputs.get("domestic_vat_rate"), "", ""],
        ["مبلغ VAT فروش داخلی", "", outputs.get("domestic_sale_vat"), ""],
        ["طلب VAT صادراتی قبلی", "", inputs.get("existing_export_vat_credit"), ""],
        ["خالص VAT قابل پرداخت یا قابل تهاتر", "", outputs.get("net_vat_position"), ""],
        ["هشدار VAT", "", "حتی در صورت تهاتر، اگر کالا معاف نباشد، VAT فروش داخلی باید در صورتحساب داخلی درج شود.", ""],
        ["هشدار دریافت VAT از خریدار", "", outputs.get("vat_buyer_payment_message"), ""],
    ])

    add_sheet("سناریوهای مالیات عملکرد", ["سناریو/مشخصه", "مبنای درآمد/سود", "نرخ مالیات", "مبلغ مالیات", "توضیح ریسک"], [
        ["اینتاکد", inputs.get("inta_code"), "", "", ""],
        ["شرح فعالیت", inputs.get("activity_description"), "", "", ""],
        ["ضریب سود فعالیت", inputs.get("estimated_profit_ratio"), "", "", ""],
        ["آیا فعالیت با پرونده مالیاتی صاحب کارت همخوان است؟", "بله" if inputs.get("inta_profile_aligned") else "خیر", "", "", ""],
        ["آیا مشاور مالیاتی تأیید کرده؟", "بله" if inputs.get("tax_advisor_confirmed") else "خیر", "", "", ""],
        ["سناریوی مالیات انتخاب‌شده", outputs.get("selected_tax_scenario_name"), "", "", ""],
        ["هزینه اداری/حسابداری", inputs.get("admin_cost"), "", "", ""],
        ["ذخیره مستقیم ریسک صاحب کارت", inputs.get("risk_reserve"), "", "", ""],
        ["A - مالیات فقط روی کارمزد", outputs.get("commission_tax_base"), inputs.get("tax_rate"), outputs.get("tax_scenario_a"), "ریسک پایین‌تر در صورت قرارداد حق‌العمل‌کاری و مستندات کامل"],
        ["B - مالیات سود برآوردی اینتاکد", outputs.get("deemed_profit"), inputs.get("tax_rate"), outputs.get("tax_scenario_b"), "ریسک متوسط؛ وابسته به پذیرش اینتاکد و مدل حسابداری"],
        ["C - مالیات سود واقعی/دلخواه", outputs.get("custom_profit"), inputs.get("tax_rate"), outputs.get("tax_scenario_c"), "ریسک بالا اگر سود واقعی و مدارک هزینه قابل دفاع نباشد"],
    ])

    add_sheet("چک‌لیست ریسک", ["کنترل", "انجام شده؟ بله/خیر", "وزن ریسک", "ریسک لحاظ‌شده", "توضیح"], checklist_frame(checklist).values.tolist())

    add_sheet("محاسبه وثیقه پیشنهادی", ["شرح ذخیره", "مبنا/نرخ", "مبلغ تومان", "توضیح"], [
        ["ذخیره ریسک VAT", inputs.get("vat_risk_reserve_percent"), outputs.get("vat_risk_reserve"), ""],
        ["مالیات سناریوی منتخب", "", outputs.get("selected_income_tax"), ""],
        ["بدترین مالیات بین سناریوها", "", outputs.get("worst_case_income_tax"), ""],
        ["درصد پوشش ریسک مالیات عملکرد", inputs.get("income_tax_risk_reserve_percent"), "", ""],
        ["ذخیره پیشنهادی مالیات عملکرد", "", outputs.get("income_tax_risk_reserve"), "بر اساس بدترین سناریوی مالیات عملکرد"],
        ["ذخیره ریسک گمرکی", inputs.get("customs_risk_reserve_percent"), outputs.get("customs_risk_reserve"), ""],
        ["ذخیره ریسک ارزی", inputs.get("fx_risk_reserve_percent"), outputs.get("fx_risk_reserve"), ""],
        ["ذخیره ریسک عدم رفع تعهد", inputs.get("export_obligation_risk_reserve_percent"), outputs.get("export_obligation_risk_reserve"), ""],
        ["ذخیره حقوقی/اداری ثابت", "", outputs.get("fixed_legal_admin_reserve"), ""],
        ["حداقل وثیقه پیشنهادی", "", outputs.get("minimum_collateral"), ""],
    ])

    add_sheet("خلاصه تصمیم مدیریتی", ["شاخص", "مقدار", "توضیح"], [
        ["ارزش کل واردات", outputs.get("import_value_total"), ""],
        ["بهای تمام‌شده کل", outputs.get("total_landed_cost"), ""],
        ["کارمزد ناخالص صاحب کارت", outputs.get("gross_commission"), ""],
        ["درصد کارمزد از ارزش واردات", outputs.get("commission_percent"), ""],
        ["نام سناریوی مالیات منتخب", outputs.get("selected_tax_scenario_name"), ""],
        ["مالیات سناریوی منتخب", outputs.get("selected_income_tax"), ""],
        ["VAT واردات", outputs.get("import_vat"), ""],
        ["VAT فروش داخلی", outputs.get("domestic_sale_vat"), ""],
        ["خالص VAT قابل پرداخت/تهاتر", outputs.get("net_vat_position"), ""],
        ["سود خالص صاحب کارت", outputs.get("net_card_owner_benefit"), ""],
        ["منفعت خالص به ازای هر دلار", outputs.get("net_benefit_per_usd"), ""],
        ["هزینه کل مجری/واردکننده واقعی", outputs.get("total_operator_cost"), ""],
        ["حداقل وثیقه پیشنهادی", outputs.get("minimum_collateral"), ""],
        ["امتیاز ریسک", outputs.get("risk_score"), ""],
        ["سطح ریسک", outputs.get("risk_level"), ""],
        ["تعداد کنترل‌های انجام‌شده", outputs.get("checked_count"), ""],
        ["تعداد کنترل‌های انجام‌نشده", outputs.get("unchecked_count"), ""],
        ["سه ریسک باز مهم", "، ".join(outputs.get("open_top_risks", [])), ""],
        ["حداقل کارمزد قابل قبول هر دلار", outputs.get("minimum_commission_per_usd"), ""],
        ["فاصله کارمزد فعلی با حداقل کارمزد پیشنهادی", outputs.get("commission_gap_per_usd"), ""],
        ["هشدار کف کارمزد", outputs.get("commission_floor_warning"), ""],
        ["تصمیم نهایی", outputs.get("final_decision"), ""],
        ["توصیه عملیاتی نهایی", outputs.get("final_recommendation"), ""],
    ])

    output = BytesIO()
    _apply_sample_style(workbook)
    workbook.save(output)
    output.seek(0)
    return output


def export_checklist_to_excel(checklist: list[Mapping[str, Any]]) -> BytesIO:
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        checklist_frame(checklist).to_excel(writer, sheet_name="چک‌لیست ریسک", index=False)
    return _style_workbook(buffer)


def export_summary_to_excel(
    summary: Mapping[str, Any], labels: Mapping[str, str] | None = None
) -> BytesIO:
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        _frame(summary, labels).to_excel(writer, sheet_name="خلاصه تصمیم مدیریتی", index=False)
    return _style_workbook(buffer)


def _write_rows(sheet, rows: list[list[Any]]) -> None:
    for row in rows:
        sheet.append(row)


def _apply_sample_style(workbook: Workbook) -> None:
    header_fill = PatternFill("solid", fgColor="173F5F")
    section_fill = PatternFill("solid", fgColor="D9EAF7")
    warning_fill = PatternFill("solid", fgColor="FFF2CC")
    alert_fill = PatternFill("solid", fgColor="F4CCCC")
    border = Border(
        left=Side(style="thin", color="CBD5E0"),
        right=Side(style="thin", color="CBD5E0"),
        top=Side(style="thin", color="CBD5E0"),
        bottom=Side(style="thin", color="CBD5E0"),
    )
    money_keywords = (
        "ارزش", "مبلغ", "قیمت", "هزینه", "بهای", "کارمزد", "ذخیره",
        "وثیقه", "مالیات", "VAT", "سود خالص", "حواله", "اسپرد",
    )

    for sheet in workbook.worksheets:
        sheet.sheet_view.rightToLeft = True
        sheet.freeze_panes = "A2"
        for row in sheet.iter_rows():
            row_text = " ".join(str(cell.value or "") for cell in row)
            for cell in row:
                cell.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
                cell.border = border
                if cell.row == 1:
                    cell.fill = header_fill
                    cell.font = Font(color="FFFFFF", bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                elif str(cell.value or "").startswith("بخش:"):
                    cell.fill = section_fill
                    cell.font = Font(bold=True, color="173F5F")
                elif "هشدار" in row_text:
                    cell.fill = warning_fill
                    cell.font = Font(bold=True, color="9C6500")
                elif "رد معامله" in row_text:
                    cell.fill = alert_fill
                    cell.font = Font(bold=True, color="9C0006")

                row_label = str(sheet.cell(cell.row, 1).value or "")
                header_label = str(sheet.cell(1, cell.column).value or "")
                if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                    if (
                        0 <= float(cell.value) <= 1
                        and ("نرخ" in header_label or "درصد" in header_label or "ضریب" in row_label)
                    ):
                        cell.number_format = '0%'
                    elif any(keyword in row_label for keyword in money_keywords):
                        cell.number_format = '#,##0'
                    elif 0 < float(cell.value) < 1:
                        cell.number_format = '0%'
                    else:
                        cell.number_format = '#,##0.00'
                elif isinstance(cell.value, str) and cell.value.startswith("="):
                    if "درصد" in row_label or "نرخ" in header_label:
                        cell.number_format = '0%'
                    elif any(keyword in str(sheet.cell(cell.row, 1).value or "") for keyword in money_keywords):
                        cell.number_format = '#,##0'

        for col_idx in range(1, sheet.max_column + 1):
            max_length = 0
            for cell in sheet[get_column_letter(col_idx)]:
                max_length = max(max_length, len(str(cell.value or "")))
            sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_length + 4, 14), 58)
        sheet.row_dimensions[1].height = 28

    if "خلاصه تصمیم مدیریتی" in workbook.sheetnames:
        summary = workbook["خلاصه تصمیم مدیریتی"]
        green_fill = PatternFill("solid", fgColor="D9EAD3")
        yellow_fill = PatternFill("solid", fgColor="FFF2CC")
        red_fill = PatternFill("solid", fgColor="F4CCCC")
        summary.conditional_formatting.add(
            "B23",
            FormulaRule(formula=['B23="قابل قبول"'], fill=green_fill),
        )
        summary.conditional_formatting.add(
            "B23",
            FormulaRule(formula=['B23="نیازمند تضمین و کنترل تکمیلی"'], fill=yellow_fill),
        )
        summary.conditional_formatting.add(
            "B23",
            FormulaRule(formula=['B23="رد معامله"'], fill=red_fill),
        )


def _add_yes_no_validation(workbook: Workbook, ranges: dict[str, list[str]]) -> None:
    for sheet_name, sheet_ranges in ranges.items():
        sheet = workbook[sheet_name]
        validation = DataValidation(type="list", formula1='"بله,خیر"', allow_blank=False)
        validation.error = "فقط یکی از گزینه‌های بله یا خیر را انتخاب کنید."
        validation.errorTitle = "ورودی نامعتبر"
        validation.prompt = "بله یا خیر را انتخاب کنید."
        validation.promptTitle = "انتخاب مجاز"
        sheet.add_data_validation(validation)
        for cell_range in sheet_ranges:
            validation.add(cell_range)


def _add_list_validation(
    workbook: Workbook, sheet_name: str, cell_range: str, options: list[str]
) -> None:
    sheet = workbook[sheet_name]
    validation = DataValidation(
        type="list",
        formula1=f'"{",".join(options)}"',
        allow_blank=False,
    )
    validation.error = "یکی از گزینه‌های مجاز را انتخاب کنید."
    validation.errorTitle = "ورودی نامعتبر"
    validation.prompt = "یک گزینه از فهرست انتخاب کنید."
    validation.promptTitle = "انتخاب مجاز"
    sheet.add_data_validation(validation)
    validation.add(cell_range)


def generate_sample_output(path: str | Path) -> Path:
    """Create a realistic Persian workbook with formulas for the sample transaction."""

    workbook = Workbook()
    workbook.remove(workbook.active)

    ws_inputs = workbook.create_sheet("ورودی‌های معامله")
    _write_rows(
        ws_inputs,
        [
            ["عنوان", "مقدار", "واحد/توضیح"],
            ["ارزش واردات ارزی", 100_000, "دلار آمریکا"],
            ["نوع ارز", "USD", ""],
            ["نرخ تبدیل", 65_000, "تومان برای هر واحد ارز"],
            ["ارزش ریالی واردات", "=B2*B4", "فرمول: ارزش ارزی × نرخ تبدیل"],
            ["HS Code", "85176290", "نمونه: تجهیزات ارتباطی"],
            ["شرح کالا", "تجهیزات شبکه و مخابرات", ""],
            ["کشور مبدأ", "چین", ""],
            ["مبلغ فاکتور فروش داخلی", 8_500_000_000, "تومان، قبل از VAT"],
            ["کارمزد هر دلار برای صاحب کارت", 2_500, "تومان"],
            ["میزان منفعت رفع تعهد ارزی", 100_000_000, "تومان"],
            ["آیا کالا معاف از VAT است؟", "خیر", ""],
            ["آیا مجوز خاص نیاز دارد؟", "بله", "مجوز ثبت سفارش/استاندارد/صمت بررسی شود"],
            ["مبلغ مورد انتظار رفع تعهد", 4_000_000_000, "تومان"],
            ["آیا VAT فروش داخلی جداگانه توسط خریدار پرداخت می‌شود؟", "بله", ""],
        ],
    )

    ws_costs = workbook.create_sheet("هزینه‌های گمرکی و واردات")
    _write_rows(
        ws_costs,
        [
            ["شرح", "مبنا/نرخ", "مبلغ تومان", "توضیح/فرمول"],
            ["قیمت کالا", "='ورودی‌های معامله'!B2", "='ورودی‌های معامله'!B2*'ورودی‌های معامله'!B4", "ارزش کالا به نرخ تبدیل"],
            ["حمل", 3_000, "=B3*'ورودی‌های معامله'!B4", "هزینه حمل خارجی به ارز"],
            ["بیمه", 500, "=B4*'ورودی‌های معامله'!B4", "هزینه بیمه خارجی به ارز"],
            ["کارمزد حواله", 0.01, "=SUM(C2:C4)*B5", "درصد از جمع کالا، حمل و بیمه"],
            ["اسپرد تبدیل ارز", 0.02, "=SUM(C2:C4)*B6", "هزینه اختلاف نرخ خرید/تسویه ارز"],
            ["تعدیل ارزش گمرکی", 0.03, "=SUM(C2:C4)*B7", "افزایش ارزش برای محاسبات گمرکی"],
            ["پایه ارزش گمرکی", "", "=SUM(C2:C4)+C7", "پایه محاسبه حقوق ورودی و سود بازرگانی"],
            ["حقوق ورودی", 0.04, "=C8*B9", "نرخ نمونه"],
            ["سود بازرگانی", 0.05, "=C8*B10", "نرخ نمونه"],
            ["سایر عوارض", "", 15_000_000, "عوارض متفرقه نمونه"],
            ["هزینه ترخیص", "", 80_000_000, "حق‌العمل، تشریفات و خدمات"],
            ["انبارداری", "", 20_000_000, ""],
            ["دموراژ", "", 50_000_000, "نمونه ریسک تأخیر"],
            ["هزینه‌های بندری", "", 30_000_000, ""],
            ["بهای تمام‌شده قبل از VAT", "", "=SUM(C2:C6)+SUM(C9:C15)", "بدون VAT واردات"],
            ["بهای تمام‌شده با VAT", "", "=C16+'محاسبات ارزش افزوده'!C4", "شامل خروج وجه VAT واردات"],
        ],
    )

    ws_vat = workbook.create_sheet("محاسبات ارزش افزوده")
    _write_rows(
        ws_vat,
        [
            ["شرح", "نرخ/ورودی", "مبلغ تومان", "توضیح"],
            ["پایه مشمول VAT واردات", "", "='هزینه‌های گمرکی و واردات'!C8+'هزینه‌های گمرکی و واردات'!C9+'هزینه‌های گمرکی و واردات'!C10+'هزینه‌های گمرکی و واردات'!C11", "پایه گمرکی + حقوق ورودی + سود بازرگانی + عوارض"],
            ["نرخ VAT واردات", 0.10, "", "نرخ نمونه"],
            ["مبلغ VAT واردات", "", '=IF(\'ورودی‌های معامله\'!B12="بله",0,C2*B3)', "اگر کالا معاف باشد صفر می‌شود"],
            ["مبلغ فاکتور فروش داخلی", "", "='ورودی‌های معامله'!B9", "مبلغ قبل از VAT"],
            ["نرخ VAT فروش داخلی", 0.10, "", "نرخ نمونه"],
            ["مبلغ VAT فروش داخلی", "", '=IF(\'ورودی‌های معامله\'!B12="بله",0,C5*B6)', "باید در صورتحساب داخلی درج شود"],
            ["طلب VAT صادراتی قبلی", "", 200_000_000, "اعتبار/طلب قبلی صاحب کارت"],
            ["خالص VAT قابل پرداخت یا قابل تهاتر", "", "=C7-C4-C8", "مثبت: پرداختنی، منفی: قابل تهاتر/اعتبار"],
            ["هشدار VAT", "", "حتی در صورت تهاتر، اگر کالا معاف نباشد، VAT فروش داخلی باید در صورتحساب داخلی درج شود.", "کنترل سامانه مؤدیان و تأیید خریدار ضروری است"],
        ],
    )

    ws_tax = workbook.create_sheet("سناریوهای مالیات عملکرد")
    _write_rows(
        ws_tax,
        [
            ["سناریو/مشخصه", "مبنای درآمد/سود", "نرخ مالیات", "مبلغ مالیات", "سود خالص صاحب کارت بعد از مالیات", "توضیح ریسک"],
            ["اینتاکد", "4690010", "", "", "", "نمونه کد فعالیت، باید با پرونده واقعی تطبیق داده شود"],
            ["شرح فعالیت", "واردات و فروش داخلی تجهیزات شبکه", "", "", "", ""],
            ["ضریب سود فعالیت", 0.10, "", "", "", "ضریب نمونه بر اساس اینتاکد"],
            ["آیا فعالیت با پرونده مالیاتی صاحب کارت همخوان است؟", "خیر", "", "", "", "عدم همخوانی، ریسک تشخیص درآمد کامل را بالا می‌برد"],
            ["آیا مشاور مالیاتی تأیید کرده؟", "بله", "", "", "", "تأیید نهایی باید مستند شود"],
            ["سناریوی مالیات انتخاب‌شده", "B - مالیات سود برآوردی اینتاکد", "", "", "", "برای خلاصه مدیریتی"],
            ["هزینه اداری/حسابداری", 30_000_000, "", "", "", "ورودی آشکار برای محاسبه سود خالص صاحب کارت"],
            ["ذخیره مستقیم ریسک صاحب کارت", 50_000_000, "", "", "", "ورودی آشکار برای کسر از منفعت خالص"],
            ["A) مالیات فقط روی کارمزد صاحب کارت", "='ورودی‌های معامله'!B2*'ورودی‌های معامله'!B10", 0.25, "=B10*C10", "=B10+'ورودی‌های معامله'!B11-D10-B8-B9", "ریسک پایین‌تر در صورت قرارداد حق‌العمل‌کاری و مستندات کامل"],
            ["B) مالیات روی سود فرضی فروش کالا بر اساس ضریب اینتاکد", "='ورودی‌های معامله'!B9*B4", 0.25, "=B11*C11", "=('ورودی‌های معامله'!B2*'ورودی‌های معامله'!B10)+'ورودی‌های معامله'!B11-D11-B8-B9", "ریسک متوسط؛ وابسته به پذیرش اینتاکد و مدل حسابداری"],
            ["C) مالیات روی سود واقعی/دلخواه", 400_000_000, 0.25, "=B12*C12", "=('ورودی‌های معامله'!B2*'ورودی‌های معامله'!B10)+'ورودی‌های معامله'!B11-D12-B8-B9", "ریسک بالا اگر سود واقعی و مدارک هزینه قابل دفاع نباشد"],
        ],
    )

    ws_risk = workbook.create_sheet("چک‌لیست ریسک")
    risk_rows = [
        ["قرارداد کتبی وجود دارد", "بله", 8, "قرارداد سه‌جانبه با تعهدات مالیاتی و گمرکی"],
        ["HS Code بررسی شده", "بله", 8, "استعلام تعرفه و حقوق ورودی انجام شده"],
        ["مجوزهای کالا بررسی شده", "خیر", 7, "برای کالای نمونه مجوز خاص لازم است"],
        ["مبلغ فاکتور داخلی مشخص است", "بله", 6, "مبنای VAT و مالیات عملکرد"],
        ["پرداخت‌کننده VAT واردات مشخص است", "بله", 6, "در قرارداد با مجری مشخص شده"],
        ["پرداخت‌کننده VAT فروش داخلی مشخص است", "بله", 6, "خریدار داخلی جداگانه پرداخت می‌کند"],
        ["صورتحساب الکترونیکی صادر می‌شود", "بله", 8, "در سامانه مؤدیان"],
        ["خریدار داخلی صورتحساب را تأیید می‌کند", "خیر", 7, "ریسک رد اعتبار و اختلاف فروش"],
        ["اینتاکد بررسی شده", "خیر", 8, "نیازمند تأیید مستند مشاور مالیاتی"],
        ["مالیات عملکرد تعیین تکلیف شده", "بله", 8, "سناریو B به‌عنوان مبنا انتخاب شده"],
        ["مسئولیت گمرکی مشخص است", "بله", 8, "مسئولیت اختلاف ارزش و تعرفه با مجری"],
        ["مسیر رفع تعهد ارزی تأیید شده", "بله", 9, "مسیر تسویه در سامانه و بانک مشخص است"],
        ["صاحب کارت کنترل سامانه جامع تجارت دارد", "بله", 7, "دسترسی اصلی نزد مالک کارت است"],
        ["مجری وثیقه معتبر می‌دهد", "بله", 10, "وثیقه بانکی/چک تضمینی"],
        ["ریسک مالیاتی با وثیقه پوشش داده شده", "بله", 9, "تا زمان قطعیت مالیات"],
        ["ریسک گمرکی با وثیقه پوشش داده شده", "بله", 9, "تا پایان ترخیص و رسیدگی احتمالی"],
        ["ریسک رد اعتبار VAT پوشش داده شده", "خیر", 9, "باید در وثیقه یا ضمانت لحاظ شود"],
        ["ریسک تأخیر ترخیص پوشش داده شده", "بله", 6, "مهلت و جریمه قراردادی تعریف شده"],
        ["ریسک دموراژ پوشش داده شده", "بله", 5, "سقف مسئولیت مجری تعیین شده"],
        ["حسابدار/حسابرس مدل را تأیید کرده", "خیر", 8, "مدل ثبت درآمد باید پیش از معامله تأیید شود"],
    ]
    ws_risk.append(["کنترل", "انجام شده؟ بله/خیر", "وزن ریسک", "ریسک لحاظ‌شده", "توضیح"])
    for idx, row in enumerate(risk_rows, start=2):
        ws_risk.append([row[0], row[1], row[2], f'=IF(B{idx}="بله",0,C{idx})', row[3]])
    ws_risk.append(["امتیاز کل ریسک", "", "", "=SUM(D2:D21)", "جمع وزن موارد انجام‌نشده"])
    ws_risk.append(["تعداد کنترل‌های انجام‌شده", "=COUNTIF(B2:B21,\"بله\")", "", "", "برای خلاصه مدیریتی"])
    ws_risk.append(["تعداد کنترل‌های انجام‌نشده", "=COUNTIF(B2:B21,\"خیر\")", "", "", "برای خلاصه مدیریتی"])

    ws_collateral = workbook.create_sheet("محاسبه وثیقه پیشنهادی")
    _write_rows(
        ws_collateral,
        [
            ["شرح ذخیره", "مبنای محاسبه", "درصد ذخیره", "مبلغ ذخیره", "توضیح"],
            ["راهنمای درصدها", "", "", "", "در ستون درصد ذخیره، عدد 1 به معنی 100% است؛ اعداد به‌صورت درصد نمایش داده می‌شوند."],
            ["ذخیره ریسک VAT", "=MAX('محاسبات ارزش افزوده'!C4,'محاسبات ارزش افزوده'!C7)", 1.00, "=B3*C3", "پوشش ریسک رد اعتبار یا عدم پرداخت"],
            ["مالیات سناریوی منتخب", '=IF(LEFT(\'سناریوهای مالیات عملکرد\'!B7,1)="A",\'سناریوهای مالیات عملکرد\'!D10,IF(LEFT(\'سناریوهای مالیات عملکرد\'!B7,1)="B",\'سناریوهای مالیات عملکرد\'!D11,\'سناریوهای مالیات عملکرد\'!D12))', "", "", "برای نمایش؛ مبنای محافظه‌کارانه ذخیره نیست"],
            ["بدترین مالیات بین سناریوها", "=MAX('سناریوهای مالیات عملکرد'!D10:D12)", "", "", "بیشترین مبلغ مالیات در سناریوهای A/B/C"],
            ["درصد پوشش ریسک مالیات عملکرد", "", 1.00, "", "عدد 1 یعنی 100% پوشش"],
            ["ذخیره پیشنهادی مالیات عملکرد", "=B5", "=C6", "=B5*C6", "برای محافظت از صاحب کارت، وثیقه مالیاتی بهتر است بر اساس بدترین سناریوی مالیات عملکرد محاسبه شود، نه فقط سناریوی منتخب."],
            ["ذخیره ریسک گمرکی", "='محاسبات ارزش افزوده'!C2", 0.20, "=B8*C8", "اختلاف ارزش، تعرفه، جریمه و عوارض"],
            ["ذخیره ریسک ارزی", "=SUM('هزینه‌های گمرکی و واردات'!C2:C4)", 0.05, "=B9*C9", "نوسان نرخ و هزینه تسویه"],
            ["ذخیره ریسک عدم رفع تعهد", "='ورودی‌های معامله'!B14", 0.20, "=B10*C10", "تا قطعی شدن مسیر رفع تعهد"],
            ["ذخیره حقوقی/اداری ثابت", 100_000_000, "", "=B11", "پیگیری حقوقی، حسابداری و اداری"],
            ["حداقل وثیقه پیشنهادی", "", "", "=SUM(D3,D7:D11)", "حداقل تضمین قابل مذاکره"],
            ["هشدار وثیقه", "", "", "هیچ معامله‌ای نباید بدون وثیقه پوشش‌دهنده ریسک VAT، مالیات، گمرک و رفع تعهد ارزی انجام شود.", "کنترل قراردادی و ضمانت اجرایی لازم است"],
        ],
    )

    ws_summary = workbook.create_sheet("خلاصه تصمیم مدیریتی")
    _write_rows(
        ws_summary,
        [
            ["شاخص", "مقدار", "توضیح"],
            ["ارزش کل واردات", "='ورودی‌های معامله'!B5", "ارزش ریالی معامله"],
            ["بهای تمام‌شده کل", "='هزینه‌های گمرکی و واردات'!C17", "با VAT واردات"],
            ["کارمزد ناخالص صاحب کارت", "='سناریوهای مالیات عملکرد'!B10", "ارزش ارزی × کارمزد هر دلار"],
            ["سناریوی مالیات انتخاب‌شده", "='سناریوهای مالیات عملکرد'!B7", "گزینه کامل انتخاب‌شده"],
            ["نام سناریوی مالیات انتخاب‌شده", "=B5", "شرح سناریوی فعال"],
            ["درصد کارمزد از ارزش واردات", "=B4/B2", "کارمزد ناخالص نسبت به ارزش ریالی واردات"],
            ["VAT واردات", "='محاسبات ارزش افزوده'!C4", ""],
            ["VAT فروش داخلی", "='محاسبات ارزش افزوده'!C7", "باید در صورتحساب درج شود مگر معافیت"],
            ["خالص VAT قابل پرداخت/تهاتر", "='محاسبات ارزش افزوده'!C9", "مثبت: پرداختنی، منفی: تهاتر/اعتبار"],
            ["مالیات عملکرد انتخاب‌شده", '=IF(LEFT(\'سناریوهای مالیات عملکرد\'!B7,1)="A",\'سناریوهای مالیات عملکرد\'!D10,IF(LEFT(\'سناریوهای مالیات عملکرد\'!B7,1)="B",\'سناریوهای مالیات عملکرد\'!D11,\'سناریوهای مالیات عملکرد\'!D12))', "بر اساس سناریوی منتخب"],
            ["سود خالص صاحب کارت", '=IF(LEFT(\'سناریوهای مالیات عملکرد\'!B7,1)="A",\'سناریوهای مالیات عملکرد\'!E10,IF(LEFT(\'سناریوهای مالیات عملکرد\'!B7,1)="B",\'سناریوهای مالیات عملکرد\'!E11,\'سناریوهای مالیات عملکرد\'!E12))', "پس از مالیات، هزینه اداری و ذخیره ریسک"],
            ["سود خالص به ازای هر دلار", "=B12/'ورودی‌های معامله'!B2", "تومان به ازای هر دلار واردات"],
            ["هزینه کل مجری/واردکننده واقعی", "='هزینه‌های گمرکی و واردات'!C17+'سناریوهای مالیات عملکرد'!B10+MAX('محاسبات ارزش افزوده'!C9,0)+'سناریوهای مالیات عملکرد'!B8", "بهای تمام‌شده + کارمزد + VAT پرداختنی + هزینه اداری"],
            ["حداقل وثیقه پیشنهادی", "='محاسبه وثیقه پیشنهادی'!D12", "جمع ذخایر ریسک"],
            ["امتیاز ریسک", "='چک‌لیست ریسک'!D22", "جمع وزن کنترل‌های انجام‌نشده"],
            ["تعداد کنترل‌های انجام‌نشده", "='چک‌لیست ریسک'!B24", "موارد خیر"],
            ["تعداد کنترل‌های انجام‌شده", "='چک‌لیست ریسک'!B23", "موارد بله"],
            ["مجموع وزن ریسک‌های انجام‌نشده", "='چک‌لیست ریسک'!D22", "همان امتیاز کل ریسک"],
            ["آیا VAT فروش داخلی جداگانه توسط خریدار پرداخت می‌شود؟", "='ورودی‌های معامله'!B15", "ورودی قراردادی حساس"],
            ["پیام وضعیت دریافت VAT از خریدار", '=IF(B20="خیر","هشدار: اگر VAT فروش داخلی جداگانه از خریدار دریافت نشود، ممکن است این مبلغ از سود صاحب کارت یا مجری کسر شود.","VAT فروش داخلی به‌صورت جداگانه از خریدار دریافت می‌شود.")', "هشدار نقدینگی VAT"],
            ["سطح ریسک", '=IF(B16<=20,"کم",IF(B16<=45,"متوسط",IF(B16<=70,"زیاد","غیرقابل قبول")))', "طبقه‌بندی امتیاز ریسک"],
            ["تصمیم نهایی", '=IF(OR(B16>70,B12<0),"رد معامله",IF(B16>=46,"نیازمند تضمین و کنترل تکمیلی",IF(B16>=21,"نیازمند تضمین و کنترل تکمیلی","قابل قبول")))', "خروجی مدیریتی نهایی"],
            ["توصیه عملیاتی نهایی", '=IF(OR(B16>70,B12<0),"رد معامله؛ ریسک یا زیان خالص برای صاحب کارت بیش از حد قابل قبول است.",IF(B16>=46,"فقط در صورت دریافت وثیقه کافی، تأیید مالیاتی/گمرکی و کنترل کامل اسناد قابل بررسی است.",IF(B16>=21,"قابل بررسی با اخذ تضمین، قرارداد دقیق، تعیین تکلیف VAT، اینتاکد و فاکتور داخلی.","از نظر مدیریتی قابل قبول است، مشروط به تأیید نهایی حسابدار، مشاور مالیاتی و کنترل اسناد.")))', "متن قابل استفاده برای تصمیم مدیریتی"],
            ["بخش: ریسک‌های باز مهم", "", "سه کنترل انجام‌نشده با بیشترین وزن ریسک"],
            ["ریسک باز مهم ۱", '=IF(B17=0,"همه کنترل‌های اصلی انجام شده‌اند.",INDEX(SORTBY(FILTER(\'چک‌لیست ریسک\'!A2:A21,\'چک‌لیست ریسک\'!B2:B21="خیر"),FILTER(\'چک‌لیست ریسک\'!C2:C21,\'چک‌لیست ریسک\'!B2:B21="خیر"),-1),1))', "بالاترین وزن"],
            ["ریسک باز مهم ۲", '=IF(B17<2,"",INDEX(SORTBY(FILTER(\'چک‌لیست ریسک\'!A2:A21,\'چک‌لیست ریسک\'!B2:B21="خیر"),FILTER(\'چک‌لیست ریسک\'!C2:C21,\'چک‌لیست ریسک\'!B2:B21="خیر"),-1),2))', "دومین وزن"],
            ["ریسک باز مهم ۳", '=IF(B17<3,"",INDEX(SORTBY(FILTER(\'چک‌لیست ریسک\'!A2:A21,\'چک‌لیست ریسک\'!B2:B21="خیر"),FILTER(\'چک‌لیست ریسک\'!C2:C21,\'چک‌لیست ریسک\'!B2:B21="خیر"),-1),3))', "سومین وزن"],
            ["بخش: تحلیل حداقل کارمزد قابل قبول", "", "تحلیل پوشش هزینه‌ها و ریسک مستقیم صاحب کارت"],
            ["کارمزد فعلی هر دلار", "='ورودی‌های معامله'!B10", "تومان برای هر دلار"],
            ["کارمزد ناخالص کل", "=B4", "جمع کارمزد ناخالص"],
            ["کارمزد خالص پس از مالیات و هزینه‌های مستقیم", "=B12-'ورودی‌های معامله'!B11", "بدون منفعت رفع تعهد ارزی"],
            ["منفعت خالص به ازای هر دلار", "=B13", "تومان برای هر دلار"],
            ["حداقل کارمزد پیشنهادی هر دلار برای پوشش مالیات، هزینه اداری و ذخیره ریسک مستقیم صاحب کارت", "=(B11+'سناریوهای مالیات عملکرد'!B8+'سناریوهای مالیات عملکرد'!B9)/'ورودی‌های معامله'!B2", "فرمول پیشنهادی مدیریتی"],
            ["فاصله کارمزد فعلی با حداقل کارمزد پیشنهادی", "=B30-B34", "مثبت یعنی کارمزد فعلی بالاتر از حداقل است"],
            ["هشدار کف کارمزد", '=IF(B35<0,"کارمزد فعلی برای پوشش هزینه‌ها و ریسک‌های مستقیم صاحب کارت کافی نیست.","کارمزد فعلی هزینه‌ها و ریسک‌های مستقیم صاحب کارت را پوشش می‌دهد.")', "کنترل کف کارمزد"],
            ["هشدار VAT", "", "حتی در صورت تهاتر، VAT فروش داخلی باید در صورتحساب داخلی درج شود مگر کالا معاف باشد."],
            ["هشدار کنترل سامانه", "", "صاحب کارت نباید کنترل دسترسی سامانه جامع تجارت را از دست بدهد."],
            ["هشدار وثیقه", "", "هیچ معامله‌ای بدون وثیقه پوشش‌دهنده VAT، مالیات، گمرک و رفع تعهد ارزی انجام نشود."],
        ],
    )

    _add_yes_no_validation(
        workbook,
        {
            "ورودی‌های معامله": ["B12:B13", "B15"],
            "سناریوهای مالیات عملکرد": ["B5:B6"],
            "چک‌لیست ریسک": ["B2:B21"],
        },
    )
    _add_list_validation(
        workbook,
        "سناریوهای مالیات عملکرد",
        "B7",
        [
            "A - مالیات فقط روی کارمزد",
            "B - مالیات سود برآوردی اینتاکد",
            "C - مالیات سود واقعی/دلخواه",
        ],
    )
    _apply_sample_style(workbook)
    destination = Path(path)
    workbook.save(destination)
    return destination
